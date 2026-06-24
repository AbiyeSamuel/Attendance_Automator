import streamlit as st
import pandas as pd
import datetime
import io
import subprocess
import sys
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# --- Logging setup ---
logging.basicConfig(filename='app.log', level=logging.WARNING,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Auto-install missing dependencies
try:
    import openpyxl
except ImportError:
    st.warning("📦 Installing missing 'openpyxl'...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    st.success("✅ openpyxl installed. Please rerun the app.")
    st.stop()

# ══════════════════════════════════════════════════════════════
# 1. PAGE CONFIG & CSS
# ══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Attendance Automator Pro",
    page_icon="🏫",
    layout="wide"
)

st.markdown("""
<style>
    h1,h2,h3 { font-family:'Segoe UI',Tahoma,Geneva,Verdana,sans-serif; }
    h1 { font-weight:800; }
    [data-testid="stFileUploadDropzone"] {
        border:2px dashed #3b82f6; border-radius:12px; padding:20px;
        background-color:rgba(59,130,246,0.05);
    }
    [data-testid="baseButton-secondary"] {
        border:2px solid #3b82f6; color:#3b82f6; border-radius:8px;
        font-weight:bold; transition:all 0.3s ease;
    }
    [data-testid="baseButton-secondary"]:hover {
        background-color:#3b82f6; color:white;
        box-shadow:0 4px 10px rgba(59,130,246,0.3);
    }
    [data-testid="baseButton-primary"] {
        background-color:#10b981; color:white; border-radius:8px;
        font-weight:bold; border:none; transition:all 0.3s ease;
    }
    [data-testid="baseButton-primary"]:hover {
        background-color:#059669;
        box-shadow:0 4px 10px rgba(16,185,129,0.3);
    }
    div[data-testid="metric-container"] {
        background-color: rgba(59,130,246,0.06);
        border: 1px solid rgba(59,130,246,0.15);
        border-radius: 10px; padding: 12px 16px;
    }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# 2. HOW TO USE (collapsible)
# ══════════════════════════════════════════════════════════════
with st.expander("📖 How to use this app", expanded=False):
    st.markdown("""
    **1. Prepare your files**  
    - **Biometric logs**: `.dat` files with 3 columns: `ID  YYYY-MM-DD  HH:MM:SS`  
    - **Student names**: CSV or Excel with columns `Attendance_ID` and `Full_Name`  

    **2. Upload both file types** (use the sidebar to adjust time rules if needed)  

    **3. Click 'Generate Pro Analytics Report'**  

    **4. Download** the full Excel report or a cleaned daily CSV.  

    *Need help? Check that IDs in the logs match the `Attendance_ID` column exactly.*
    """)

# ══════════════════════════════════════════════════════════════
# 3. SIDEBAR — CONFIGURABLE TIME RULES
# ══════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("# 🏫 Attendance Automator Pro")
    st.header("⏱️ Attendance Time Rules")
    st.caption("Adjust cutoff times to match your class schedule.")

    st.subheader("🌅 Morning Session")
    m_present_cutoff = st.time_input("Present if clocked in by", datetime.time(10, 30), key="mpc")

    st.subheader("🌇 Afternoon Session")
    a_present_cutoff = st.time_input("Present if clocked in by", datetime.time(14, 30), key="apc")

    st.divider()
    st.subheader("🗜️ Time Compression")
    st.caption("Compress variable clock-in windows into a standard grading window.")
    use_compression = st.checkbox("Enable time compression", value=True)
    m_src_end = st.time_input("Morning source window end", datetime.time(11, 30), key="mse")
    a_src_end = st.time_input("Afternoon source window end", datetime.time(14, 50), key="ase")
    target_minutes = st.number_input("Target window (minutes)", min_value=10, max_value=60, value=30, step=5)

    st.divider()
    report_title = st.text_input("Report Title", value="SHIP DESIGN AND CONSTRUCTION ATTENDANCE REPORT")

# ══════════════════════════════════════════════════════════════
# 4. TIME COMPRESSION ENGINE
# ══════════════════════════════════════════════════════════════
def scale_time(time_str: str) -> str:
    if not use_compression:
        return time_str
    try:
        t = datetime.datetime.strptime(time_str, '%H:%M:%S')
        tgt_secs = target_minutes * 60

        m_start = datetime.datetime.strptime('10:00:00', '%H:%M:%S')
        m_end = datetime.datetime.strptime(m_src_end.strftime('%H:%M:%S'), '%H:%M:%S')

        a_start = datetime.datetime.strptime('14:00:00', '%H:%M:%S')
        a_end = datetime.datetime.strptime(a_src_end.strftime('%H:%M:%S'), '%H:%M:%S')

        if m_start <= t <= m_end:
            src_dur = (m_end - m_start).total_seconds()
            new_offset = datetime.timedelta(seconds=int((t - m_start).total_seconds() * (tgt_secs / src_dur)))
            return (m_start + new_offset).strftime('%H:%M:%S')

        if a_start <= t <= a_end:
            src_dur = (a_end - a_start).total_seconds()
            new_offset = datetime.timedelta(seconds=int((t - a_start).total_seconds() * (tgt_secs / src_dur)))
            return (a_start + new_offset).strftime('%H:%M:%S')

        return time_str
    except Exception as e:
        logging.warning(f"Time compression error for '{time_str}': {e}")
        return time_str

# ══════════════════════════════════════════════════════════════
# 5. GRADING ENGINE (Arrival Based Only)
# ══════════════════════════════════════════════════════════════
def grade_session(compressed_time_str: str, session: str) -> tuple:
    try:
        t = datetime.datetime.strptime(compressed_time_str, '%H:%M:%S').time()
        if session == 'Morning':
            if t <= m_present_cutoff: return 'PRESENT', 0.5
            return 'LATE', 0.25
        else:  # Afternoon Session
            if t <= a_present_cutoff: return 'PRESENT', 0.5
            return 'LATE', 0.25
    except Exception as e:
        logging.warning(f"Grading error for '{compressed_time_str}' ({session}): {e}")
        return 'ABSENT', 0.0

# ══════════════════════════════════════════════════════════════
# 6. PROFESSIONAL EXCEL STYLING & CHARTS ENGINE
# ══════════════════════════════════════════════════════════════
def style_workbook(output_bytes: bytes, has_unknown_ids: bool, generation_time: str) -> bytes:
    BLUE_DARK, BLUE_MID, BLUE_LIGHT = "1E3A5F", "2563EB", "DBEAFE"
    GREEN_DARK, GREEN_LIGHT = "065F46", "D1FAE5"
    AMBER_DARK, AMBER_LIGHT = "92400E", "FEF3C7"
    RED_DARK, RED_LIGHT = "991B1B", "FEE2E2"
    GREY_LIGHT, WHITE, PURPLE = "F8FAFC", "FFFFFF", "7C3AED"

    thin = Side(style='thin', color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = load_workbook(filename=io.BytesIO(output_bytes))

    def style_header_row(ws, row=1, color=BLUE_DARK):
        for cell in ws[row]:
            cell.font = Font(name="Segoe UI", bold=True, color=WHITE, size=10)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    # ── Sheet : Executive_Summary ────────────────────────────
    ws = wb["Executive_Summary"]
    ws.sheet_properties.tabColor = "1E3A5F"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_area = f'A1:B{ws.max_row}'
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    ws.insert_rows(1, 3)
    ws.merge_cells("A1:B3")
    c = ws["A1"]
    c.value = f"🏫  {report_title}"
    c.font = Font(name="Segoe UI", size=16, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 48

    for col, val in enumerate(["Metric", "Value"], start=1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.font = Font(name="Segoe UI", bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=BLUE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ts_row = ws.max_row + 2
    ws.cell(row=ts_row, column=1, value="Report generated on:").font = Font(italic=True)
    ws.cell(row=ts_row, column=2, value=generation_time).font = Font(italic=True)

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=2):
        for i, cell in enumerate(row):
            cell.font = Font(name="Segoe UI", size=11)
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=GREY_LIGHT if cell.row % 2 == 1 else WHITE)
            cell.alignment = Alignment(horizontal="center" if i == 1 else "left", vertical="center")
        ws.row_dimensions[row[0].row].height = 22

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 26

    # ── Extract Summary Data for Insights ──
    try:
        avg_att_str = ws.cell(row=9, column=2).value
        avg_att = float(str(avg_att_str).replace('%', '')) if avg_att_str else 0.0
        total_students = int(ws.cell(row=7, column=2).value)
    except:
        avg_att, total_students = 0.0, 0

    # ── Sheet : Delegate_Review ───────────────────────────────
    ws2 = wb["Delegate_Review"]
    ws2.sheet_properties.tabColor = "2563EB"
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "C2"

    headers = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
    week_cols = [i + 1 for i, h in enumerate(headers) if h and str(h).startswith('Week_')]
    overall_col = next((i + 1 for i, h in enumerate(headers) if h and 'Overall' in str(h)), None)
    remarks_col = next((i + 1 for i, h in enumerate(headers) if h and 'Remark' in str(h)), None)
    risk_col = next((i + 1 for i, h in enumerate(headers) if h and 'Risk' in str(h)), None)

    style_header_row(ws2)

    if overall_col:
        ws2.conditional_formatting.add(
            f'{get_column_letter(overall_col)}2:{get_column_letter(overall_col)}{ws2.max_row}',
            CellIsRule(operator='lessThan', formula=['50'], fill=PatternFill("solid", fgColor="FEE2E2"))
        )

    if remarks_col:
        dv = DataValidation(type="list",
                            formula1='"Medical,Travel,Family Emergency,Other"',
                            allow_blank=True)
        dv.error = "Please select a valid reason"
        dv.errorTitle = "Invalid Remark"
        ws2.add_data_validation(dv)
        dv.add(f'{get_column_letter(remarks_col)}2:{get_column_letter(remarks_col)}{ws2.max_row}')

    for row_idx, row in enumerate(ws2.iter_rows(min_row=2, max_row=ws2.max_row), start=2):
        bg = GREY_LIGHT if row_idx % 2 == 0 else WHITE
        for cell in row:
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center" if cell.column > 2 else "left", vertical="center",
                                       wrap_text=(cell.column == remarks_col))
            raw_val = str(cell.value) if cell.value is not None else ""

            def colour_pct(cell, raw_val, size=10, bold=False):
                try:
                    pct = float(raw_val.replace('%', ''))
                    if pct >= 75:
                        col = GREEN_DARK
                    elif pct >= 50:
                        col = AMBER_DARK
                    else:
                        col = RED_DARK
                    cell.font = Font(name="Segoe UI", size=size, bold=bold, color=col)
                except ValueError:
                    cell.font = Font(name="Segoe UI", size=size)

            if cell.column == overall_col:
                colour_pct(cell, raw_val, bold=True)
            elif cell.column in week_cols:
                try:
                    pct = float(raw_val.replace('%', ''))
                    if pct >= 75:
                        cell.fill, cell.font = PatternFill("solid", fgColor=GREEN_LIGHT), Font(name="Segoe UI", size=9,
                                                                                               color=GREEN_DARK)
                    elif pct >= 50:
                        cell.fill, cell.font = PatternFill("solid", fgColor=AMBER_LIGHT), Font(name="Segoe UI", size=9,
                                                                                               color=AMBER_DARK)
                    else:
                        cell.fill, cell.font = PatternFill("solid", fgColor=RED_LIGHT), Font(name="Segoe UI", size=9,
                                                                                             color=RED_DARK)
                except ValueError:
                    cell.font = Font(name="Segoe UI", size=9)
            elif cell.column == risk_col:
                risk_colors = {"High Risk": (RED_DARK, RED_LIGHT), "Moderate Risk": (AMBER_DARK, AMBER_LIGHT),
                               "Low Risk": (GREEN_DARK, GREEN_LIGHT)}
                if raw_val in risk_colors:
                    fc, bg_c = risk_colors[raw_val]
                    cell.font, cell.fill = Font(name="Segoe UI", size=9, bold=True, color=fc), PatternFill("solid",
                                                                                                           fgColor=bg_c)
                else:
                    cell.font = Font(name="Segoe UI", size=9)
            elif cell.column == remarks_col:
                cell.font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
            else:
                cell.font = Font(name="Segoe UI", size=10)

    fixed_widths = {1: 6, 2: 32}
    for col_idx in range(1, ws2.max_column + 1):
        w = fixed_widths.get(col_idx, 14 if col_idx in week_cols else (32 if col_idx == remarks_col else 16))
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[1].height = 32

    # ── Sheet : Detailed_Grid ─────────────────────────────────
    ws3 = wb["Detailed_Grid"]
    ws3.sheet_properties.tabColor = "DBEAFE"
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "C3"

    status_style = {"PRESENT": (GREEN_LIGHT, GREEN_DARK), "LATE": (AMBER_LIGHT, AMBER_DARK),
                    "ABSENT": (RED_LIGHT, RED_DARK)}

    for row_idx, row in enumerate(ws3.iter_rows(), start=1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            val = str(cell.value) if cell.value is not None else ""

            if row_idx == 1:
                cell.fill, cell.font = PatternFill("solid", fgColor=BLUE_DARK), Font(name="Segoe UI", size=10,
                                                                                     bold=True, color=WHITE)
            elif row_idx == 2:
                cell.fill, cell.font = PatternFill("solid", fgColor=BLUE_LIGHT), Font(name="Segoe UI", size=9,
                                                                                      bold=True, color=BLUE_MID)
            elif val in status_style:
                bg_c, fc = status_style[val]
                cell.fill, cell.font = PatternFill("solid", fgColor=bg_c), Font(name="Segoe UI", size=9, bold=True,
                                                                                color=fc)
            else:
                cell.fill, cell.font = PatternFill("solid", fgColor=GREY_LIGHT if row_idx % 2 == 0 else WHITE), Font(
                    name="Segoe UI", size=10)

    ws3.column_dimensions["A"].width, ws3.column_dimensions["B"].width = 8, 30
    for col in range(3, ws3.max_column + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 12
    ws3.row_dimensions[1].height, ws3.row_dimensions[2].height = 28, 22

    # ── Sheet : Risk_Analysis ─────────────────────────────────
    risk_counts = {"High Risk": 0, "Moderate Risk": 0, "Low Risk": 0}
    if "Risk_Analysis" in wb.sheetnames:
        ws_r = wb["Risk_Analysis"]
        ws_r.sheet_properties.tabColor = PURPLE
        ws_r.sheet_view.showGridLines = False
        style_header_row(ws_r, color=PURPLE)
        risk_colors = {"High Risk": (RED_DARK, RED_LIGHT), "Moderate Risk": (AMBER_DARK, AMBER_LIGHT),
                       "Low Risk": (GREEN_DARK, GREEN_LIGHT)}

        for row_idx, row in enumerate(ws_r.iter_rows(min_row=2, max_row=ws_r.max_row), start=2):
            bg = GREY_LIGHT if row_idx % 2 == 0 else WHITE
            for cell in row:
                cell.border, cell.alignment = border, Alignment(horizontal="center" if cell.column > 2 else "left",
                                                                vertical="center")
                cell.fill, cell.font = PatternFill("solid", fgColor=bg), Font(name="Segoe UI", size=10)
                val = str(cell.value) if cell.value else ""

                if cell.column == 4 and val in risk_colors:
                    fc, bg_c = risk_colors[val]
                    cell.font, cell.fill = Font(name="Segoe UI", size=10, bold=True, color=fc), PatternFill("solid",
                                                                                                            fgColor=bg_c)
                    if val in risk_counts: risk_counts[val] += 1

        ws_r.column_dimensions["A"].width, ws_r.column_dimensions["B"].width = 6, 32
        ws_r.column_dimensions["C"].width, ws_r.column_dimensions["D"].width = 18, 18

        start_row = ws_r.max_row + 3
        ws_r.cell(row=start_row, column=1, value="Risk Level").font = Font(bold=True)
        ws_r.cell(row=start_row, column=2, value="Count").font = Font(bold=True)
        for i, (risk_lbl, count) in enumerate(risk_counts.items(), 1):
            ws_r.cell(row=start_row + i, column=1, value=risk_lbl)
            ws_r.cell(row=start_row + i, column=2, value=count)

        pc = PieChart()
        pc.title = "Risk Level Distribution"
        data = Reference(ws_r, min_col=2, min_row=start_row, max_row=start_row + len(risk_counts))
        cats = Reference(ws_r, min_col=1, min_row=start_row + 1, max_row=start_row + len(risk_counts))
        pc.add_data(data, titles_from_data=True)
        pc.set_categories(cats)
        ws_r.add_chart(pc, "F2")

    # ── Sheet : Daily_Trend ───────────────────────────────────
    if "Daily_Trend" in wb.sheetnames:
        ws_t = wb["Daily_Trend"]
        ws_t.sheet_properties.tabColor = GREEN_DARK
        ws_t.sheet_view.showGridLines = False
        style_header_row(ws_t, color=BLUE_DARK)
        for row_idx, row in enumerate(ws_t.iter_rows(min_row=2, max_row=ws_t.max_row), start=2):
            for cell in row:
                cell.border, cell.fill, cell.font, cell.alignment = border, PatternFill("solid",
                                                                                        fgColor=GREY_LIGHT if row_idx % 2 == 0 else WHITE), Font(
                    name="Segoe UI", size=10), Alignment(horizontal="center", vertical="center")
        ws_t.column_dimensions["A"].width, ws_t.column_dimensions["B"].width = 18, 24

        lc = LineChart()
        lc.title, lc.y_axis.title, lc.x_axis.title = "Daily Attendance Rate Trend", "Attendance Rate (%)", "Date"
        lc.style, lc.width, lc.height = 10, 22, 14
        data = Reference(ws_t, min_col=2, min_row=1, max_row=ws_t.max_row)
        cats = Reference(ws_t, min_col=1, min_row=2, max_row=ws_t.max_row)
        lc.add_data(data, titles_from_data=True)
        lc.set_categories(cats)
        ws_t.add_chart(lc, "D2")

    # ── Sheet : Session_Heatmap ───────────────────────────────
    if "Session_Heatmap" in wb.sheetnames:
        ws_h = wb["Session_Heatmap"]
        ws_h.sheet_properties.tabColor = AMBER_DARK
        ws_h.sheet_view.showGridLines = False
        style_header_row(ws_h, color=BLUE_DARK)
        for row_idx, row in enumerate(ws_h.iter_rows(min_row=2, max_row=ws_h.max_row), start=2):
            for cell in row:
                cell.border, cell.fill, cell.font, cell.alignment = border, PatternFill("solid",
                                                                                        fgColor=GREY_LIGHT if row_idx % 2 == 0 else WHITE), Font(
                    name="Segoe UI", size=10), Alignment(horizontal="center", vertical="center")
        ws_h.column_dimensions["A"].width, ws_h.column_dimensions["B"].width, ws_h.column_dimensions[
            "C"].width = 18, 24, 24

        bc = BarChart()
        bc.type, bc.grouping = "col", "stacked"
        bc.title, bc.y_axis.title, bc.x_axis.title = "Session Performance: Morning vs Afternoon", "Attendance Rate (%)", "Date"
        bc.style, bc.width, bc.height = 10, 22, 14
        data2 = Reference(ws_h, min_col=2, min_row=1, max_row=ws_h.max_row, max_col=3)
        cats2 = Reference(ws_h, min_col=1, min_row=2, max_row=ws_h.max_row)
        bc.add_data(data2, titles_from_data=True)
        bc.set_categories(cats2)
        ws_h.add_chart(bc, "E2")

    # ── Sheet : Distribution_Bands ────────────────────────────
    if "Distribution_Bands" in wb.sheetnames:
        ws_d = wb["Distribution_Bands"]
        ws_d.sheet_properties.tabColor = GREEN_LIGHT
        ws_d.sheet_view.showGridLines = False
        style_header_row(ws_d, color=BLUE_DARK)
        for row_idx, row in enumerate(ws_d.iter_rows(min_row=2, max_row=ws_d.max_row), start=2):
            for cell in row:
                cell.border, cell.fill, cell.font, cell.alignment = border, PatternFill("solid",
                                                                                        fgColor=GREY_LIGHT if row_idx % 2 == 0 else WHITE), Font(
                    name="Segoe UI", size=10), Alignment(horizontal="center", vertical="center")
        ws_d.column_dimensions["A"].width, ws_d.column_dimensions["B"].width = 18, 22

        bc2 = BarChart()
        bc2.title, bc2.x_axis.title, bc2.y_axis.title = "Attendance Distribution by Band", "Attendance Range", "Number of Delegates"
        bc2.style, bc2.width, bc2.height = 10, 20, 14
        data3 = Reference(ws_d, min_col=2, min_row=1, max_row=ws_d.max_row)
        cats3 = Reference(ws_d, min_col=1, min_row=2, max_row=ws_d.max_row)
        bc2.add_data(data3, titles_from_data=True)
        bc2.set_categories(cats3)
        ws_d.add_chart(bc2, "D2")

    # ── Sheet : Unknown_IDs ────────────────────
    if has_unknown_ids and "Unknown_IDs" in wb.sheetnames:
        ws_u = wb["Unknown_IDs"]
        ws_u.sheet_properties.tabColor = PURPLE
        ws_u.sheet_view.showGridLines = False
        style_header_row(ws_u, color=PURPLE)
        for row in ws_u.iter_rows(min_row=2):
            for cell in row:
                cell.font, cell.alignment, cell.border = Font(name="Segoe UI", size=10), Alignment(
                    horizontal="center"), border
        ws_u.column_dimensions["A"].width, ws_u.column_dimensions["B"].width, ws_u.column_dimensions[
            "C"].width = 18, 18, 50

    # ── Sheet : Automated Insights ────────────────────────────
    ws_i = wb.create_sheet("Automated_Insights")
    ws_i.sheet_properties.tabColor = BLUE_MID
    ws_i.sheet_view.showGridLines = False
    ws_i.column_dimensions['A'].width = 90

    ws_i['A1'] = "AUTOMATED INSIGHTS & RECOMMENDATIONS"
    ws_i['A1'].font = Font(name="Segoe UI", bold=True, size=14, color=WHITE)
    ws_i['A1'].fill = PatternFill("solid", fgColor=BLUE_DARK)

    high_risk, mod_risk = risk_counts.get("High Risk", 0), risk_counts.get("Moderate Risk", 0)

    insights = [
        f"• Overall attendance rate is {avg_att:.1f}%. {'Excellent engagement.' if avg_att >= 75 else 'Below expected threshold (75%). Immediate intervention recommended.'}",
        f"• {high_risk} students ({high_risk / total_students * 100:.1f}%) are at HIGH RISK (attendance < 50%).",
        f"• {mod_risk} students are at MODERATE RISK (50% - 74%). Targeted follow-up needed."
    ]

    row = 3
    for ins in insights:
        ws_i.cell(row=row, column=1, value=ins).font = Font(name="Segoe UI", size=11)
        row += 1

    ws_i.cell(row=row + 1, column=1, value="RECOMMENDED ACTION ITEMS:").font = Font(name="Segoe UI", bold=True, size=12)
    row += 2
    recos = [
        "• Send formal warning to all High-Risk students.",
        "• Implement a buddy system or study group for Moderate-Risk students.",
        "• Review session timings or material if specific days show consistently low turnout.",
        "• Congratulate and reward students maintaining >90% attendance to boost morale."
    ]
    for rec in recos:
        ws_i.cell(row=row, column=1, value=rec).font = Font(name="Segoe UI", size=11)
        row += 1

    styled = io.BytesIO()
    wb.save(styled)
    styled.seek(0)
    return styled.read()

# ══════════════════════════════════════════════════════════════
# 7. APP HEADER
# ══════════════════════════════════════════════════════════════
st.title("🏫 Weekly Attendance Automator")
st.markdown(
    "Upload biometric `.dat` logs and a Names file → get a **fully graded, "
    "analytics-enriched Excel report** with charts, risk flags, weekly breakdowns, "
    "and a cleaned daily summary CSV."
)
st.divider()

# ══════════════════════════════════════════════════════════════
# 8. UPLOAD ZONES
# ══════════════════════════════════════════════════════════════
col1, col2 = st.columns(2)
with col1:
    uploaded_files = st.file_uploader(
        "📂 1. Biometric Logs (.dat) — multiple files allowed",
        type=["dat"], accept_multiple_files=True
    )
with col2:
    names_file = st.file_uploader(
        "📝 2. Student Names (.csv or .xlsx)",
        type=["csv", "xlsx"]
    )

if uploaded_files and names_file:

    raw_records = []
    malformed_lines = 0
    total_lines = 0
    progress_bar = st.progress(0)
    status_text = st.empty()

    for i, f in enumerate(uploaded_files):
        status_text.text(f"Parsing file {i+1} of {len(uploaded_files)}: {f.name}")
        lines = f.getvalue().decode("utf-8", errors='replace').splitlines()
        total_lines += len(lines)
        for line in lines:
            parts = line.strip().split()
            if len(parts) >= 3:
                try:
                    datetime.datetime.strptime(parts[1].strip(), '%Y-%m-%d')
                    datetime.datetime.strptime(parts[2].strip(), '%H:%M:%S')
                    raw_records.append({
                        "ID": parts[0].strip(),
                        "Date": parts[1].strip(),
                        "Time_Raw": parts[2].strip(),
                        "Time_Grading": scale_time(parts[2].strip()),
                    })
                except ValueError:
                    malformed_lines += 1
            else:
                malformed_lines += 1
        progress_bar.progress((i + 1) / len(uploaded_files))

    status_text.text("Parsing complete.")
    if malformed_lines > 0:
        st.warning(f"⚠️ {malformed_lines} malformed line(s) skipped out of {total_lines} total.")

    if not raw_records:
        st.error("No valid records found in the uploaded .dat files.")
        st.stop()

    raw_df = pd.DataFrame(raw_records)
    raw_df['Session'] = raw_df['Time_Raw'].apply(
        lambda x: 'Morning' if int(x.split(':')[0]) < 13 else 'Afternoon'
    )

    first_scan = (
        raw_df.sort_values('Time_Raw')
        .drop_duplicates(subset=['ID', 'Date', 'Session'], keep='first')
        .reset_index(drop=True)
    )

    st.markdown("<br>", unsafe_allow_html=True)

    if st.button("🚀 Generate Pro Analytics Report", use_container_width=True):
        with st.spinner("Processing scans, grading attendance, building analytics..."):
            try:
                names_df = (
                    pd.read_csv(names_file) if names_file.name.endswith('.csv')
                    else pd.read_excel(names_file)
                )
                if not {'Attendance_ID', 'Full_Name'}.issubset(names_df.columns):
                    st.error("Names file must have columns: `Attendance_ID` and `Full_Name`.")
                    st.stop()

                dup_ids = names_df[names_df['Attendance_ID'].duplicated(keep=False)]
                if not dup_ids.empty:
                    st.warning(f"⚠️ Duplicate `Attendance_ID` found: {', '.join(dup_ids['Attendance_ID'].astype(str).unique())}. Only the last occurrence per ID will be used.")

                names_df['Attendance_ID'] = (
                    names_df['Attendance_ID'].astype(str)
                    .str.split('.').str[0].str.strip()
                )
                id_to_name = dict(zip(names_df['Attendance_ID'], names_df['Full_Name']))

                first_scan['Full_Name'] = first_scan['ID'].map(id_to_name)
                unknown_df = first_scan[first_scan['Full_Name'].isna()].copy()
                known_first = first_scan[first_scan['Full_Name'].notna()].copy()

                known_first[['Status', 'Points']] = known_first.apply(
                    lambda row: pd.Series(grade_session(row['Time_Grading'], row['Session'])),
                    axis=1
                )
                graded_df = known_first[['ID', 'Full_Name', 'Date', 'Session', 'Status', 'Points', 'Time_Raw']].copy()

                all_dates = sorted(graded_df['Date'].unique())
                total_days = len(all_dates)

                date_to_week = {}
                week_labels = []
                for d in all_dates:
                    iso_w = pd.Timestamp(d).isocalendar().week
                    if iso_w not in [w for _, w in date_to_week.values()]:
                        week_labels.append(f"Week_{len(week_labels) + 1}")
                    date_to_week[d] = (week_labels[-1], iso_w)
                week_order = sorted(set(v[0] for v in date_to_week.values()))

                all_students = names_df[['Attendance_ID', 'Full_Name']].copy()
                all_students['Attendance_ID'] = all_students['Attendance_ID'].astype(str).str.strip()
                total_delegates = len(all_students)

                pivot = graded_df.pivot_table(
                    index=['ID', 'Full_Name'],
                    columns=['Date', 'Session'],
                    values='Status',
                    aggfunc='first'
                )
                full_index = pd.MultiIndex.from_arrays(
                    [all_students['Attendance_ID'], all_students['Full_Name']],
                    names=['ID', 'Full_Name']
                )
                full_columns = pd.MultiIndex.from_product(
                    [all_dates, ['Morning', 'Afternoon']],
                    names=['Date', 'Session']
                )
                pivot = pivot.reindex(index=full_index, columns=full_columns, fill_value='ABSENT')
                pivot = pivot.reset_index()
                pivot['_s'] = pd.to_numeric(pivot['ID'], errors='coerce')
                pivot = pivot.sort_values('_s').drop(columns='_s').set_index(['ID', 'Full_Name'])

                earned = (
                    graded_df.groupby('ID')['Points'].sum()
                    .reset_index().rename(columns={'Points': 'Total Points Earned'})
                )

                graded_df['Week_Label'] = graded_df['Date'].map(lambda d: date_to_week[d][0])
                days_per_week = (
                    graded_df[['Date', 'Week_Label']].drop_duplicates()
                    .groupby('Week_Label')['Date'].count().rename('days_in_week')
                )
                weekly_pts = graded_df.groupby(['ID', 'Week_Label'])['Points'].sum().reset_index()
                weekly_pts = weekly_pts.merge(days_per_week, on='Week_Label')
                weekly_pts['Week_Pct'] = (weekly_pts['Points'] / weekly_pts['days_in_week']).clip(0, 1)
                weekly_pivot = (
                    weekly_pts.pivot(index='ID', columns='Week_Label', values='Week_Pct')
                    .reindex(columns=week_order)
                )

                total_possible = float(total_days)
                review = all_students.rename(columns={'Attendance_ID': 'ID'}).copy()
                review['ID'] = review['ID'].astype(str)
                review = review.merge(earned, on='ID', how='left')
                review = review.merge(weekly_pivot.reset_index(), on='ID', how='left')
                review['Total Points Earned'] = review['Total Points Earned'].fillna(0.0)
                for wk in week_order:
                    review[wk] = review.get(wk, pd.Series(0.0, index=review.index)).fillna(0.0)

                review['Total Possible'] = total_possible
                review['Overall_Performance (%)'] = (
                        (review['Total Points Earned'] / total_possible * 100).round(1).astype(str) + '%'
                )
                for wk in week_order:
                    review[wk] = (review[wk] * 100).round(1).astype(str) + '%'

                def risk_level(pts):
                    pct = (pts / total_possible * 100) if total_possible > 0 else 0
                    if pct >= 75:   return "Low Risk"
                    if pct >= 50:   return "Moderate Risk"
                    return "High Risk"

                review['Risk Level'] = review['Total Points Earned'].apply(risk_level)
                review['Remarks / Reason for Low Attendance'] = ''

                review['_s'] = pd.to_numeric(review['ID'], errors='coerce')
                review = review.sort_values('_s').drop(columns='_s').reset_index(drop=True)
                review_cols = (
                        ['ID', 'Full_Name', 'Total Points Earned', 'Total Possible']
                        + week_order
                        + ['Overall_Performance (%)', 'Risk Level',
                           'Remarks / Reason for Low Attendance']
                )
                review = review[review_cols]

                daily_att = (
                        graded_df.groupby('Date')['Points'].sum() /
                        (total_delegates * 1.0) * 100
                )
                trend_df = pd.DataFrame({
                    'Date': daily_att.index,
                    'Attendance Rate (%)': daily_att.values.round(1)
                })

                def session_rate(session):
                    s = graded_df[graded_df['Session'] == session].groupby('Date')['Points'].sum()
                    return (s / (total_delegates * 0.5) * 100).reindex(all_dates).fillna(0).round(1)

                heatmap_df = pd.DataFrame({
                    'Date': all_dates,
                    'Morning Attendance (%)': session_rate('Morning').values,
                    'Afternoon Attendance (%)': session_rate('Afternoon').values,
                })

                pct_series = review['Total Points Earned'].apply(
                    lambda p: (p / total_possible * 100) if total_possible > 0 else 0
                )
                bins = [0, 50, 75, 90, 101]
                labels = ['0–50%', '50–75%', '75–90%', '90–100%']
                dist_df = (
                    pd.cut(pct_series, bins=bins, labels=labels, right=False)
                    .value_counts()
                    .rename_axis('Attendance Band')
                    .reset_index(name='Number of Delegates')
                    .sort_values('Attendance Band')
                )

                risk_df = review[['ID', 'Full_Name', 'Overall_Performance (%)', 'Risk Level']].copy()

                morning_merge = graded_df[graded_df['Session'] == 'Morning'][
                    ['ID', 'Date', 'Time_Raw', 'Status', 'Points']].rename(
                    columns={'Time_Raw': 'Morning_Arrival', 'Status': 'Morning_Status', 'Points': 'Morning_Points'})
                afternoon_merge = graded_df[graded_df['Session'] == 'Afternoon'][
                    ['ID', 'Date', 'Time_Raw', 'Status', 'Points']].rename(
                    columns={'Time_Raw': 'Afternoon_Arrival', 'Status': 'Afternoon_Status',
                             'Points': 'Afternoon_Points'})

                all_date_ids = pd.MultiIndex.from_product(
                    [all_students['Attendance_ID'], all_dates], names=['ID', 'Date']
                ).to_frame(index=False)
                cleaned = (
                    all_date_ids
                    .merge(morning_merge, on=['ID', 'Date'], how='left')
                    .merge(afternoon_merge, on=['ID', 'Date'], how='left')
                )
                cleaned['Full_Name'] = cleaned['ID'].map(id_to_name)
                cleaned['Morning_Status'] = cleaned['Morning_Status'].fillna('ABSENT')
                cleaned['Afternoon_Status'] = cleaned['Afternoon_Status'].fillna('ABSENT')
                cleaned['Morning_Points'] = cleaned['Morning_Points'].fillna(0)
                cleaned['Afternoon_Points'] = cleaned['Afternoon_Points'].fillna(0)
                cleaned['Daily_Points'] = cleaned['Morning_Points'] + cleaned['Afternoon_Points']
                cleaned = cleaned[['ID', 'Full_Name', 'Date',
                                   'Morning_Arrival', 'Morning_Status',
                                   'Afternoon_Arrival', 'Afternoon_Status',
                                   'Daily_Points']].sort_values(['Date', 'ID'])

                present_count = int((graded_df['Status'] == 'PRESENT').sum())
                late_count = int((graded_df['Status'] == 'LATE').sum())
                absent_count = int(total_delegates * total_days * 2 - len(graded_df))
                unknown_count = len(unknown_df['ID'].unique()) if not unknown_df.empty else 0
                avg_attendance = (
                        review['Total Points Earned'].sum() /
                        (total_delegates * total_possible) * 100
                ) if total_delegates > 0 else 0.0
                high_risk_count = int((review['Risk Level'] == 'High Risk').sum())

                summary_df = pd.DataFrame({
                    "Metric": [
                        "Report Title",
                        "Total Class Days Logged",
                        "Total Registered Delegates",
                        "Total Weeks Covered",
                        "Average Overall Attendance Rate",
                        "Total PRESENT Records",
                        "Total LATE Records",
                        "Total Missed Sessions",
                        "High Risk Delegates (< 50%)",
                        "Unknown Biometric IDs Detected",
                    ],
                    "Value": [
                        report_title,
                        total_days,
                        total_delegates,
                        len(week_order),
                        f"{avg_attendance:.1f}%",
                        present_count,
                        late_count,
                        absent_count,
                        high_risk_count,
                        unknown_count,
                    ]
                })

                gen_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    summary_df.to_excel(writer, sheet_name='Executive_Summary', index=False)
                    review.to_excel(writer, sheet_name='Delegate_Review', index=False)
                    pivot.to_excel(writer, sheet_name='Detailed_Grid')
                    risk_df.to_excel(writer, sheet_name='Risk_Analysis', index=False)
                    trend_df.to_excel(writer, sheet_name='Daily_Trend', index=False)
                    heatmap_df.to_excel(writer, sheet_name='Session_Heatmap', index=False)
                    dist_df.to_excel(writer, sheet_name='Distribution_Bands', index=False)
                    if not unknown_df.empty:
                        (unknown_df.groupby('ID')
                         .agg(Scan_Count=('Date', 'count'),
                              Dates_Seen=('Date', lambda x: ', '.join(sorted(x.unique()))))
                         .reset_index()
                         .to_excel(writer, sheet_name='Unknown_IDs', index=False))

                styled_bytes = style_workbook(
                    output.getvalue(), has_unknown_ids=not unknown_df.empty, generation_time=gen_time
                )

                st.success("✅ Report generated successfully!")
                st.balloons()

                st.subheader("📊 Executive Summary")
                r1c1, r1c2, r1c3, r1c4 = st.columns(4)
                r1c1.metric("📅 Class Days", total_days)
                r1c2.metric("👥 Registered Delegates", total_delegates)
                r1c3.metric("📈 Avg Attendance", f"{avg_attendance:.1f}%")
                r1c4.metric("📆 Weeks", len(week_order))

                r2c1, r2c2, r2c3 = st.columns(3)
                r2c1.metric("✅ PRESENT", present_count)
                r2c2.metric("⚠️ LATE", late_count)
                r2c3.metric("❌ Missed Sessions", absent_count)

                r3c1, r3c2, r3c3 = st.columns(3)
                r3c1.metric("🔴 High Risk Delegates", high_risk_count)
                r3c2.metric("🟡 Moderate Risk",
                            int((review['Risk Level'] == 'Moderate Risk').sum()))
                r3c3.metric("🟢 Low Risk",
                            int((review['Risk Level'] == 'Low Risk').sum()))

                st.subheader("📋 Delegate Review (interactive)")
                review_display = review.copy()
                review_display['Overall_Performance_Num'] = review_display['Overall_Performance (%)'].str.rstrip('%').astype(float)
                def color_risk(val):
                    if val == 'High Risk':
                        return 'background-color: #FEE2E2; color: #991B1B; font-weight: bold'
                    elif val == 'Moderate Risk':
                        return 'background-color: #FEF3C7; color: #92400E; font-weight: bold'
                    elif val == 'Low Risk':
                        return 'background-color: #D1FAE5; color: #065F46; font-weight: bold'
                    return ''
                styled_review = review_display.style.map(color_risk, subset=['Risk Level'])\
                    .format({'Overall_Performance_Num': '{:.1f}%'})\
                    .hide(axis=1, subset=['Overall_Performance_Num'])
                st.dataframe(styled_review, use_container_width=True, height=400)

                st.subheader("🔍 Student Detail View")
                student_ids = review['ID'].tolist()
                selected_student = st.selectbox("Select a student to view their daily attendance", student_ids)
                if selected_student:
                    student_data = graded_df[graded_df['ID'] == selected_student].copy()
                    if not student_data.empty:
                        student_daily = student_data.groupby('Date')['Points'].sum().reindex(all_dates, fill_value=0).reset_index()
                        student_daily.columns = ['Date', 'Points']
                        student_daily['Possible'] = 1.0
                        st.line_chart(student_daily.set_index('Date')['Points'])
                        st.caption(f"Daily points for {id_to_name.get(selected_student, selected_student)}. 1.0 = both sessions present.")

                with st.expander("📈 View Daily Attendance Trend", expanded=False):
                    st.line_chart(trend_df.set_index('Date')['Attendance Rate (%)'])

                with st.expander("🌅🌇 Morning vs Afternoon Breakdown", expanded=False):
                    st.bar_chart(heatmap_df.set_index('Date'))

                with st.expander("🏷️ Attendance Distribution Bands", expanded=False):
                    st.bar_chart(dist_df.set_index('Attendance Band'))

                if not unknown_df.empty:
                    st.warning(
                        f"⚠️ **{unknown_count} unknown biometric ID(s)** found "
                        f"({', '.join(sorted(unknown_df['ID'].unique()))}) — "
                        f"not in your Names list. See the **Unknown_IDs** sheet."
                    )

                st.info(
                    "💡 The **Remarks / Reason for Low Attendance** column in "
                    "**Delegate_Review** is blank — fill it in Excel before "
                    "submitting to management."
                )

                st.subheader("📥 Download Center")
                dl1, dl2 = st.columns(2)
                with dl1:
                    st.download_button(
                        label="📊 Download Full Analytics Report (.xlsx)",
                        data=styled_bytes,
                        file_name="Weekly_Attendance_Pro_Report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
                with dl2:
                    st.download_button(
                        label="📄 Download Cleaned Daily Summary (.csv)",
                        data=cleaned.to_csv(index=False).encode('utf-8'),
                        file_name="Cleaned_Daily_Summary.csv",
                        mime="text/csv",
                        type="secondary",
                        use_container_width=True
                    )

                # ── UPDATED: Feedback via email (mailto link) ──
                st.divider()
                st.subheader("📝 Feedback")
                st.markdown("We'd love to hear your thoughts. Click the button below to send feedback directly to **samuel.abiye@cmotd.org** via your email client.")
                mailto_link = f"mailto:samuel.abiye@cmotd.org?subject=Attendance%20Automator%20Feedback&body=Please%20enter%20your%20feedback%20below:%0A%0A"
                st.link_button("✉️ Send Feedback via Email", mailto_link)

            except KeyError as e:
                st.error(
                    f"Column not found: **{e}**. "
                    f"Ensure your Names file has columns: `Attendance_ID` and `Full_Name`."
                )
            except Exception as e:
                st.error(f"Unexpected error: {e}")
                logging.exception("Unhandled exception during report generation")
                raise

            #Will be updated